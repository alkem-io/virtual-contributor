"""File text extraction: PDF, DOCX, XLSX."""

from __future__ import annotations

import logging
from io import BytesIO

logger = logging.getLogger(__name__)


def parse_pdf(content: bytes) -> str:
    """Extract text from PDF bytes."""
    from pypdf import PdfReader
    reader = PdfReader(BytesIO(content))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def parse_docx(content: bytes) -> str:
    """Extract text from DOCX bytes."""
    from docx import Document as DocxDocument
    doc = DocxDocument(BytesIO(content))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_xlsx(content: bytes) -> str:
    """Extract text from XLSX bytes."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts)


def parse_file(content: bytes, filename: str) -> str | None:
    """Parse a file based on its extension. Returns None if unsupported."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    parsers = {
        "pdf": parse_pdf,
        "docx": parse_docx,
        "xlsx": parse_xlsx,
    }

    parser = parsers.get(ext)
    if parser is None:
        logger.warning("Unsupported file format: %s", ext)
        return None

    try:
        return parser(content)
    except Exception as exc:
        logger.error("Failed to parse %s: %s", filename, exc)
        return None
